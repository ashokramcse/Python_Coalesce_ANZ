'''
Created on 20-Jun-2017

@author: BALASUBRAMANIAM
'''
import calendar
from openpyxl import Workbook
from openpyxl import load_workbook
from pip._vendor.distlib.compat import raw_input
from openpyxl.cell import get_column_letter
'''
dirName="G:/Local Disk/Python/ANZ_Day4"
fileName=raw_input("Enter file Name")
filePath=dirName+"/"+fileName+".xlsx"
wb=Workbook()
i=1
for month in calendar.month_name:
    print(month)
    if not (len((str(month)))==0):
       wb.create_sheet(month+"_2017", i)
    i+=1
wb.save(filename=filePath)
'''
fileName="G:/Local Disk/Python/ANZ_Day4/Test.xlsx"
#load the workbook
wb = load_workbook(fileName, read_only=False)
sheetNames=wb.get_sheet_names()
print(sheetNames)
sheet=wb.get_sheet_by_name(sheetNames[1])
#print(sheet)
for row in range(10, 20):
    for col in range(3, 27):
        sheet.cell(column=col, row=row, value="%s"
                      % get_column_letter(col))
        
wb.save(fileName)



