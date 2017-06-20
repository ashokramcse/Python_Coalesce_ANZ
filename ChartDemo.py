'''
Created on 20-Jun-2017

@author: BALASUBRAMANIAM
'''
import os
from openpyxl import Workbook

from openpyxl import  load_workbook
from openpyxl.chart import(AreaChart,
    Reference,
    Series,
    BarChart3D)
wb=load_workbook('G:/Local Disk/Python/ANZ_Day4/LinkData.xlsx', read_only=False,data_only=True)
sheet=wb.get_sheet_by_name("Sheet1")
outerList=[]
for cell in range(1,sheet.max_row):
    innerList=[]
    for col in range(3,5):
#        print(cell,sheet.cell(row=cell,column=col).value)
        innerList.append(sheet.cell(row=cell,column=col).value)
    outerList.append(innerList)

print(outerList)

chart = BarChart3D()
chart.title = "Link Utilization"
chart.style = 13
chart.x_axis.title = 'Time'
chart.y_axis.title = 'bandwidth'

cats = Reference(sheet, min_col=3, min_row=1, max_row=7)
data = Reference(sheet, min_col=4, min_row=1, max_col=4, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

sheet.add_chart(chart, "G10")
print(os.getcwd())
wb.save('G:/Local Disk/Python/ANZ_Day4/LinkData.xlsx')

